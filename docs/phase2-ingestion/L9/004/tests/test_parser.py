import json, subprocess, sys
from pathlib import Path
import importlib.util
P = Path(__file__).parents[1]
spec = importlib.util.spec_from_file_location('l9', P / 'parser.py')
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)
RAW = Path(__file__).parents[3] / "data/wgc/raw/gdt/GDT_Tables_Q2'26_EN.xlsx"

def test_real_workbook_components_and_blocked(tmp_path):
    rows = mod.parse_workbook(RAW, stale_after_days=1000)
    assert len(rows) > 0
    assert {r['component'] for r in rows} == {'jewellery_demand_tonnes', 'bar_and_coin_demand_tonnes', 'gross_bullion_imports_tonnes', 'net_bullion_imports_tonnes'}
    out = tmp_path / 'out.csv'
    subprocess.run([sys.executable, str(P / 'parser.py'), '--input', str(tmp_path / 'missing.xlsx'), '--output', str(out)], check=True)
    assert json.loads(out.with_suffix('.status.json').read_text())['status'] == 'BLOCKED'

def test_quarter_end_and_complete_period_fallback(tmp_path):
    rows = mod.parse_workbook(RAW, stale_after_days=1000)
    q2 = [r for r in rows if r['observation_period'] == "Q2'26"]
    assert q2 and all((r['observation_date'] == '2026-06-30' and r['observation_period_type'] == 'quarterly' for r in q2))
    annual = [r for r in rows if r['observation_period'] == '2010']
    assert annual and all((r['observation_period_type'] == 'annual' for r in annual))
    prior = tmp_path / 'prior.csv'
    import csv
    with prior.open('w', newline='') as h:
        w = csv.DictWriter(h, fieldnames=mod.FIELDS)
        w.writeheader()
        w.writerows(rows)
    carried = mod.carry_forward(prior)
    latest = max((r['observation_date'] for r in rows))
    assert len(carried) == sum((r['observation_date'] == latest for r in rows))
    assert {r['component'] for r in carried} == {r['component'] for r in rows if r['observation_date'] == latest}

def test_cli_failure_with_prior_is_stale_and_recovery_cleans_status(tmp_path):
    rows = mod.parse_workbook(RAW, stale_after_days=1000)
    prior = tmp_path / 'prior.csv'
    out = tmp_path / 'out.csv'
    import csv
    with prior.open('w', newline='') as h:
        w = csv.DictWriter(h, fieldnames=mod.FIELDS)
        w.writeheader()
        w.writerows(rows)
    status = out.with_suffix('.status.json')
    status.write_text('{"status":"BLOCKED"}')
    result = subprocess.run([sys.executable, str(P / 'parser.py'), '--input', str(tmp_path / 'missing.xlsx'), '--prior', str(prior), '--output', str(out)], capture_output=True, text=True)
    assert result.returncode == 0 and out.exists() and (not status.exists())
    with out.open(newline='') as h:
        carried = list(csv.DictReader(h))
    latest = max((r['observation_date'] for r in rows))
    assert len(carried) == sum((r['observation_date'] == latest for r in rows))
    assert {r['component'] for r in carried} == {r['component'] for r in rows if r['observation_date'] == latest}
    assert all((r['availability_status'] == 'STALE' for r in carried))
    result = subprocess.run([sys.executable, str(P / 'parser.py'), '--input', str(RAW), '--output', str(out)], capture_output=True, text=True)
    assert result.returncode == 0 and out.exists() and (not status.exists())

def test_malformed_cell_is_rejected(tmp_path):
    from openpyxl import Workbook
    b = Workbook()
    s = b.active
    s.title = 'Jewellery'
    s.cell(1, 1, 'Title')
    s.cell(2, 3, 2025)
    s.cell(3, 2, 'India')
    s.cell(3, 3, 'bad')
    for name, label in [('Bar and Coin', 'India'), ('India Supply', 'Gross Bullion Imports'), ('India Supply', 'Net Bullion Imports')]:
        sh = b[name] if name in b.sheetnames else b.create_sheet(name)
        sh.cell(1, 1, 'Title')
        sh.cell(2, 3, 2025)
        sh.cell(3, 2, label)
        sh.cell(3, 3, 1)
    bad = tmp_path / 'bad.xlsx'
    b.save(bad)
    import pytest
    with pytest.raises(ValueError):
        mod.parse_workbook(bad)

def test_large_component_is_flagged(tmp_path):
    from openpyxl import Workbook
    b = Workbook()
    s = b.active
    s.title = 'Jewellery'
    s.cell(1, 1, 'Title')
    s.cell(2, 3, '2025')
    s.cell(3, 2, 'India')
    s.cell(3, 3, 20000)
    sh = b.create_sheet('Bar and Coin')
    sh.cell(1, 1, 'Title')
    sh.cell(2, 3, '2025')
    sh.cell(3, 2, 'India')
    sh.cell(3, 3, 1)
    sh = b.create_sheet('India Supply')
    sh.cell(1, 1, 'Title')
    sh.cell(2, 3, '2025')
    sh.cell(3, 2, 'Gross Bullion Imports')
    sh.cell(3, 3, 1)
    sh.cell(4, 2, 'Net Bullion Imports')
    sh.cell(4, 3, 1)
    p = tmp_path / 'large.xlsx'
    b.save(p)
    assert any((r['validation_status'] == 'FLAG' for r in mod.parse_workbook(p)))

def test_conflicting_duplicate_component_is_rejected(tmp_path):
    from openpyxl import Workbook
    b = Workbook()
    s = b.active
    s.title = 'Jewellery'
    s.cell(1, 1, 'Title')
    s.cell(2, 3, '2025')
    s.cell(3, 2, 'India')
    s.cell(3, 3, 1)
    s.cell(4, 2, 'India')
    s.cell(4, 3, 2)
    sh = b.create_sheet('Bar and Coin')
    sh.cell(1, 1, 'Title')
    sh.cell(2, 3, '2025')
    sh.cell(3, 2, 'India')
    sh.cell(3, 3, 1)
    sh = b.create_sheet('India Supply')
    sh.cell(1, 1, 'Title')
    sh.cell(2, 3, '2025')
    sh.cell(3, 2, 'Gross Bullion Imports')
    sh.cell(3, 3, 1)
    sh.cell(4, 2, 'Net Bullion Imports')
    sh.cell(4, 3, 1)
    p = tmp_path / 'dup.xlsx'
    b.save(p)
    import pytest
    with pytest.raises(ValueError):
        mod.parse_workbook(p)
