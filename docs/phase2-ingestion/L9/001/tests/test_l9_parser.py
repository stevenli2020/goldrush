import csv
import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from jsonschema import Draft7Validator, FormatChecker
from openpyxl import Workbook
_spec = importlib.util.spec_from_file_location('l9001_parser', Path(__file__).parents[1] / 'parser.py')
_module = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_module)
parse_workbook = _module.parse_workbook
refresh_availability = _module.refresh_availability

class L9001Tests(unittest.TestCase):

    def workbook(self, root, rows, *, sheet_name='Chinese premiums-discounts', title='Chinese Premium/Discount (US$/oz) - 5 day moving average'):
        path = Path(root) / 'gold-premiums.xlsx'
        book = Workbook()
        sheet = book.active
        sheet.title = sheet_name
        sheet.append(['Disclaimer'])
        sheet.append(['Methodology: https://www.gold.org'])
        sheet.append([title])
        sheet.append(['Source: World Gold Council'])
        sheet.append(['Values are published series data'])
        for row in rows:
            sheet.append(row)
        india = book.create_sheet('Indian premiums-discounts')
        india.append(['Indian Premium/Discount (US$/oz) - 5 day moving average'])
        india.append(['2026-08-14', 999])
        book.create_sheet('Disclaimer')
        book.save(path)
        return path

    def manifest(self, root, workbook, **overrides):
        path = Path(root) / 'manifest.json'
        payload = {'target': 'gold_premiums', 'page_url': 'https://www.gold.org/goldhub/data/gold-premium', 'download_url': 'https://www.gold.org/download/file/11657/gold-premiums.xlsx', 'filename': 'gold-premiums.xlsx', 'raw_path': str(workbook), 'size_bytes': workbook.stat().st_size, 'downloaded_at': '2026-08-24T00:00:00+00:00', 'http_status': 200, 'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
        payload.update(overrides)
        path.write_text(json.dumps(payload), encoding='utf-8')
        return path

    def test_real_layout_chinese_selection_title_and_schema(self):
        with tempfile.TemporaryDirectory() as root:
            workbook = self.workbook(root, [['2026-08-14', 12.5], ['2026-08-13', -3.25]])
            rows = parse_workbook(workbook, manifest_path=self.manifest(root, workbook), stale_after_days=1000)
            self.assertEqual([row['observation_date'] for row in rows], ['2026-08-13', '2026-08-14'])
            self.assertEqual(rows[0]['premium_discount_usd_per_oz'], -3.25)
            self.assertEqual(rows[0]['smoothing_method'], '5-day moving average as published by WGC')
            schema = json.loads((Path(__file__).parents[1] / 'schema.json').read_text())
            self.assertEqual(list(Draft7Validator(schema, format_checker=FormatChecker()).iter_errors(rows[0])), [])

    def test_invalid_values_duplicates_title_sheet_and_manifest(self):
        with tempfile.TemporaryDirectory() as root:
            for values in ([['bad', 1]], [['2026-08-14', 'bad']], [['2026-08-14', float('nan')]]):
                with self.assertRaises(ValueError):
                    parse_workbook(self.workbook(root, values))
            with self.assertRaises(ValueError):
                parse_workbook(self.workbook(root, [['2026-08-14', 1], ['2026-08-14', 2]]))
            with self.assertRaises(ValueError):
                parse_workbook(self.workbook(root, [['2026-08-14', 1]], sheet_name='Indian premiums-discounts'))
            with self.assertRaises(ValueError):
                parse_workbook(self.workbook(root, [['2026-08-14', 1]], title='Chinese series'))
            workbook = self.workbook(root, [['2026-08-14', 1]])
            with self.assertRaises(ValueError):
                parse_workbook(workbook, manifest_path=self.manifest(root, workbook, target='wrong'))

    def test_revision_fallback_blocked_recovery_and_schema(self):
        with tempfile.TemporaryDirectory() as root:
            root = Path(root)
            workbook = self.workbook(root, [['2026-08-14', 1]])
            prior_rows = parse_workbook(workbook, manifest_path=self.manifest(root, workbook))
            prior = root / 'prior.csv'
            with prior.open('w', newline='') as handle:
                writer = csv.DictWriter(handle, fieldnames=list(prior_rows[0]))
                writer.writeheader()
                writer.writerows(prior_rows)
            changed = self.workbook(root, [['2026-08-14', -2]])
            self.assertTrue(parse_workbook(changed, manifest_path=self.manifest(root, changed), prior_path=prior)[0]['is_revised'])
            script = Path(__file__).parents[1] / 'parser.py'
            output = root / 'out.csv'
            result = subprocess.run([sys.executable, str(script), '--workbook', str(root / 'missing.xlsx'), '--prior', str(prior), '--output', str(output)], capture_output=True, text=True)
            self.assertEqual(result.returncode, 0)
            self.assertFalse(output.with_suffix('.status.json').exists())
            with output.open(newline='') as handle:
                carried = next(csv.DictReader(handle))
            self.assertEqual(carried['availability_status'], 'STALE')
            typed = dict(carried)
            typed['premium_discount_usd_per_oz'] = float(typed['premium_discount_usd_per_oz'])
            typed['is_revised'] = typed['is_revised'].lower() == 'true'
            typed['prior_value'] = None
            typed['raw_size_bytes'] = int(typed['raw_size_bytes'])
            typed['http_status'] = int(typed['http_status'])
            schema = json.loads((Path(__file__).parents[1] / 'schema.json').read_text())
            self.assertEqual(list(Draft7Validator(schema, format_checker=FormatChecker()).iter_errors(typed)), [])
            blocked = root / 'blocked.csv'
            subprocess.run([sys.executable, str(script), '--workbook', str(root / 'missing.xlsx'), '--output', str(blocked)], check=True)
            self.assertTrue(blocked.with_suffix('.status.json').exists())
            subprocess.run([sys.executable, str(script), '--workbook', str(changed), '--manifest', str(self.manifest(root, changed)), '--output', str(blocked)], check=True)
            self.assertFalse(blocked.with_suffix('.status.json').exists())

    def test_manual_manifest_uses_shared_fields(self):
        with tempfile.TemporaryDirectory() as root:
            root = Path(root)
            workbook = self.workbook(root, [['2026-08-14', 1]])
            manifest = root / 'manifest.json'
            subprocess.run([sys.executable, str(Path(__file__).parents[1] / 'create_manifest.py'), str(workbook), str(manifest)], check=True, capture_output=True, text=True)
            payload = json.loads(manifest.read_text())
            self.assertEqual(payload['target'], 'gold_premiums')
            self.assertEqual(payload['filename'], 'gold-premiums.xlsx')
            self.assertEqual(payload['size_bytes'], workbook.stat().st_size)
            self.assertIsNone(payload['http_status'])

    def test_refresh_status_marks_unchanged_old_output_stale(self):
        with tempfile.TemporaryDirectory() as root:
            root = Path(root)
            workbook = self.workbook(root, [['2000-01-01', 1]])
            rows = parse_workbook(workbook, manifest_path=self.manifest(root, workbook))
            output = root / 'output.csv'
            with output.open('w', newline='') as handle:
                writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
                writer.writeheader()
                writer.writerows(rows)
            refreshed = refresh_availability(output, stale_after_days=10)
            self.assertEqual(refreshed[0]['availability_status'], 'STALE')
if __name__ == '__main__':
    unittest.main()
