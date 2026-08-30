import json, subprocess, sys
from pathlib import Path
import importlib.util
P = Path(__file__).parents[1]
spec = importlib.util.spec_from_file_location('l5', P / 'parser.py')
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)
RAW = Path(__file__).parents[3] / 'data/wgc/raw/central-bank/Changes_latest_as_of_Aug2026_IFS.xlsx'

def test_real_workbook_canonical_negative_rows_and_blocked(tmp_path):
    rows = mod.parse_workbook(RAW, stale_after_days=1000)
    assert rows and all((r['official_sector_net_reduction_tonnes'] > 0 for r in rows))
    assert all((not r['country_entity'].endswith('*') for r in rows))
    out = tmp_path / 'out.csv'
    subprocess.run([sys.executable, str(P / 'parser.py'), '--input', str(tmp_path / 'missing.xlsx'), '--output', str(out)], check=True)
    assert json.loads(out.with_suffix('.status.json').read_text())['status'] == 'BLOCKED'

def test_complete_period_fallback_and_malformed_cell(tmp_path):
    rows = mod.parse_workbook(RAW, stale_after_days=1000)
    prior = tmp_path / 'prior.csv'
    import csv
    with prior.open('w', newline='') as h:
        w = csv.DictWriter(h, fieldnames=mod.FIELDS)
        w.writeheader()
        w.writerows(rows)
    carried = mod.carry_forward(prior)
    latest = max((r['observation_date'] for r in rows))
    assert len(carried) == sum((r['observation_date'] == latest for r in rows))
    assert all((r['availability_status'] == 'STALE' for r in carried))
    from openpyxl import Workbook
    b = Workbook()
    s = b.active
    s.title = 'Monthly'
    s.append(['', '', 'Country', '2026-01-01'])
    s.append(['', '', 'X', 'bad'])
    bad = tmp_path / 'bad.xlsx'
    b.save(bad)
    import pytest
    with pytest.raises(ValueError):
        mod.parse_workbook(bad)

def test_large_reduction_is_flagged(tmp_path):
    from openpyxl import Workbook
    b = Workbook()
    s = b.active
    s.title = 'Monthly'
    s.append(['', 'Country', '', '2026-01-01'])
    s.append(['', 'X', '', -20000])
    p = tmp_path / 'large.xlsx'
    b.save(p)
    assert mod.parse_workbook(p)[0]['validation_status'] == 'FLAG'

def test_conflicting_duplicate_country_is_rejected(tmp_path):
    from openpyxl import Workbook
    b = Workbook()
    s = b.active
    s.title = 'Monthly'
    s.append(['', 'Country', '', '2026-01-01'])
    s.append(['', 'X', '', -1])
    s.append(['', 'X', '', -2])
    p = tmp_path / 'dup.xlsx'
    b.save(p)
    import pytest
    with pytest.raises(ValueError):
        mod.parse_workbook(p)

def test_cli_failure_with_prior_is_stale_and_recovery_cleans_status(tmp_path):
    rows = mod.parse_workbook(RAW, stale_after_days=1000)
    prior = tmp_path / 'prior.csv'
    out = tmp_path / 'out.csv'
    import csv
    with prior.open('w', newline='') as h:
        w = csv.DictWriter(h, fieldnames=mod.FIELDS)
        w.writeheader()
        w.writerows(rows)
    status = out.with_suffix('.status.json')
    status.write_text('{"status":"BLOCKED"}')
    result = subprocess.run([sys.executable, str(P / 'parser.py'), '--input', str(tmp_path / 'missing.xlsx'), '--prior', str(prior), '--output', str(out)], capture_output=True, text=True)
    assert result.returncode == 0 and out.exists() and (not status.exists())
    with out.open(newline='') as h:
        carried = list(csv.DictReader(h))
    latest = max((r['observation_date'] for r in rows))
    assert len(carried) == sum((r['observation_date'] == latest for r in rows))
    assert all((r['availability_status'] == 'STALE' for r in carried))
    result = subprocess.run([sys.executable, str(P / 'parser.py'), '--input', str(RAW), '--output', str(out)], capture_output=True, text=True)
    assert result.returncode == 0 and out.exists() and (not status.exists())
