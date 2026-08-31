import json
import sys
from pathlib import Path
import pytest
from openpyxl import Workbook
SCRIPTS_DIR = Path(__file__).resolve().parent.parent / 'scripts'
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))
from parse_gold_recycling import GoldRecyclingCollector, parse_quarter_label

@pytest.fixture
def config_file(tmp_path):
    cfg_path = tmp_path / 'config.yaml'
    cfg_content = f'''\nvariable_id: "L0-006"\nunit: "tonnes"\nfrequency: "quarterly"\nsource:\n  target_sheets: ["Supply"]\n  search_keywords: ["recycled gold", "recycling"]\n  header_regex: "^(Q[1-4][\\\\s'’]?\\\\d{{2,4}}|\\\\d{{4}}\\\\s?Q[1-4])$"\npaths:\n  shared_raw_workbook: "{tmp_path}/cached.xlsx"\n  seed_csv: "{tmp_path}/seed.csv"\n  processed_output: "{tmp_path}/output.json"\nvalidation:\n  hard_min_value: 0.0\n  warning_min_value: 150.0\n  warning_max_value: 600.0\n'''
    cfg_path.write_text(cfg_content)
    return str(cfg_path)

def test_quarter_header_regex_parsing():
    assert parse_quarter_label("Q1'26") == '2026-03-31'
    assert parse_quarter_label('Q1 2026') == '2026-03-31'
    assert parse_quarter_label('2026 Q2') == '2026-06-30'
    assert parse_quarter_label("Q3'25") == '2025-09-30'
    assert parse_quarter_label('Invalid') is None

def test_ambiguous_row_matching_raises_error(config_file, tmp_path):
    wb_path = tmp_path / 'ambiguous.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Supply'
    ws.cell(row=1, column=1, value='Category')
    ws.cell(row=1, column=2, value="Q1'26")
    ws.cell(row=2, column=1, value='Recycled gold')
    ws.cell(row=2, column=2, value=300)
    ws.cell(row=3, column=1, value='Total Recycling')
    ws.cell(row=3, column=2, value=300)
    wb.save(wb_path)
    collector = GoldRecyclingCollector(config_file)
    with pytest.raises(ValueError, match='Ambiguous matching'):
        collector.extract_from_workbook(wb_path)

def test_negative_value_hard_failure(config_file, tmp_path):
    collector = GoldRecyclingCollector(config_file)
    raw = [{'observation_date': '2026-03-31', 'value': -10.0}]
    with pytest.raises(ValueError, match='Hard Validation Failed'):
        collector.validate_and_apply_revisions(raw, tmp_path / 'nonexistent.json')

def test_seed_csv_parsing_strips_comments(config_file, tmp_path):
    seed_path = tmp_path / 'seed.csv'
    seed_path.write_text('# Header Comment line 1\n# Header Comment line 2\nobservation_date,value,unit,frequency\n2026-03-31,310.5,tonnes,quarterly\n')
    collector = GoldRecyclingCollector(config_file)
    data = collector.parse_seed_csv(seed_path)
    assert len(data) == 1
    assert data[0]['observation_date'] == '2026-03-31'
    assert data[0]['value'] == 310.5

def test_fallback_status_marking(config_file, tmp_path):
    seed_path = tmp_path / 'seed.csv'
    seed_path.write_text('observation_date,value\n2026-03-31,310.5\n')
    collector = GoldRecyclingCollector(config_file)
    res = collector.run(publication_date='2026-05-01')
    assert res['availability_status'] == 'STALE'
    assert res['data'][0]['revision_status'] == 'ORIGINAL'

def test_ambiguous_row_matching_raises_error(config_file, tmp_path):
    wb_path = tmp_path / 'ambiguous.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Supply'
    ws.cell(row=1, column=1, value='Category')
    ws.cell(row=1, column=2, value="Q1'26")
    ws.cell(row=1, column=3, value="Q2'26")
    ws.cell(row=2, column=1, value='Recycled gold')
    ws.cell(row=2, column=2, value=300)
    ws.cell(row=3, column=1, value='Total Recycling')
    ws.cell(row=3, column=2, value=300)
    wb.save(wb_path)
    collector = GoldRecyclingCollector(config_file)
    with pytest.raises(ValueError, match='Ambiguous matching'):
        collector.extract_from_workbook(wb_path)
