import hashlib
import json
import re
import yaml
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Union


def compute_sha256(file_path: Union[str, Path]) -> Optional[str]:
    """Calculates SHA-256 hash of a given source file."""
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        return None

    sha256_hash = hashlib.sha256()
    with open(path, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()


def parse_quarter_label(label: Any) -> Optional[str]:
    """
    Parses a quarter string label (e.g., 'Q1 2025', '2025-Q1', "Q1'25", '2025-03-31')
    and returns the end-of-quarter date in ISO format YYYY-MM-DD.
    Returns None for invalid or unparseable input.
    """
    if label is None:
        return None

    s = str(label).strip()
    if not s or s.lower() in ["none", "null", "invalid", "nan"]:
        return None

    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            datetime.strptime(s, "%Y-%m-%d")
            return s
        except ValueError:
            return None

    q_match = re.search(r"Q([1-4])", s, re.IGNORECASE)
    y_match = re.search(r"(\b\d{4}\b|\b\d{2}\b)", s)

    if q_match and y_match:
        q_num = int(q_match.group(1))
        year_str = y_match.group(1)
        year = (2000 + int(year_str)) if len(year_str) == 2 else int(year_str)

        quarter_end_dates = {
            1: f"{year}-03-31",
            2: f"{year}-06-30",
            3: f"{year}-09-30",
            4: f"{year}-12-31",
        }
        return quarter_end_dates[q_num]

    return None


class GoldRecyclingCollector:
    """Collector for L0-006 Gold Recycling Flow data compliant with pipeline specs."""

    def __init__(
        self,
        config_path: Optional[Union[str, Path]] = None,
        variable_id: str = "L0-006",
        seed_path: Optional[Union[str, Path]] = None,
        output_path: Optional[Union[str, Path]] = None,
    ):
        self.variable_id = variable_id
        self.config_path = Path(config_path) if config_path else None
        self.seed_path = Path(seed_path) if seed_path else Path("data/raw/seeds/l0_006_recycling_seed.csv")
        self.output_path = Path(output_path) if output_path else Path("processed/l0_006_gold_recycling_flow.json")

        if self.config_path and self.config_path.exists():
            self._load_config(self.config_path)

    def _load_config(self, config_path: Path) -> None:
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config = yaml.safe_load(f) or {}
            if "variable_id" in config:
                self.variable_id = config["variable_id"]
            if "seed_path" in config:
                self.seed_path = Path(config["seed_path"])
            if "output_path" in config:
                self.output_path = Path(config["output_path"])
        except Exception:
            pass

    def validate_and_apply_revisions(
        self,
        observations: List[Dict[str, Any]],
        existing_output_path: Optional[Union[str, Path]] = None
    ) -> List[Dict[str, Any]]:
        """
        Validates observation entries, explicitly raising ValueError with 'Hard Validation Failed' on negative values, 
        deduplicating internally, and applying revisions.
        """
        dated_records: Dict[str, Dict[str, Any]] = {}

        for obs in observations:
            date = obs.get("observation_date")
            if not date:
                continue

            try:
                val = float(obs["value"])
            except (ValueError, TypeError, KeyError):
                continue

            if val < 0:
                raise ValueError(f"Hard Validation Failed: Invalid negative observation value {val}")

            unit = obs.get("unit", "tonnes")
            freq = obs.get("frequency", "quarterly")
            rev_status = obs.get("revision_status", "ORIGINAL")

            if date in dated_records:
                prev_obs = dated_records[date]
                prev_val = prev_obs["value"]
                if val != prev_val:
                    dated_records[date] = {
                        "observation_date": date,
                        "value": val,
                        "unit": unit,
                        "frequency": freq,
                        "revision_status": "REVISED",
                        "previous_value": prev_val,
                        "revision_notes": obs.get("revision_notes", "Value revised against internal duplicate"),
                    }
            else:
                dated_records[date] = {
                    "observation_date": date,
                    "value": val,
                    "unit": unit,
                    "frequency": freq,
                    "revision_status": rev_status,
                }
                if "previous_value" in obs:
                    dated_records[date]["previous_value"] = obs["previous_value"]
                if "revision_notes" in obs:
                    dated_records[date]["revision_notes"] = obs["revision_notes"]

        if existing_output_path:
            ext_path = Path(existing_output_path)
            if ext_path.exists():
                try:
                    with open(ext_path, "r", encoding="utf-8") as f:
                        prev_payload = json.load(f)
                        prev_list = prev_payload.get("observations") or prev_payload.get("data", [])
                        for prev_obs in prev_list:
                            date = prev_obs.get("observation_date")
                            if date in dated_records:
                                current_val = dated_records[date]["value"]
                                historical_val = float(prev_obs["value"])
                                if current_val != historical_val:
                                    dated_records[date]["revision_status"] = "REVISED"
                                    dated_records[date]["previous_value"] = historical_val
                                    dated_records[date]["revision_notes"] = "Revised from previous ingestion run"
                except Exception:
                    pass

        return sorted(dated_records.values(), key=lambda x: x["observation_date"])

    def parse_seed_csv(self, file_path: Optional[Union[str, Path]] = None) -> List[Dict[str, Any]]:
        """Parses a seed CSV file returning raw observations."""
        target_path = Path(file_path) if file_path else self.seed_path
        observations: List[Dict[str, Any]] = []

        if not target_path.exists():
            return observations

        with open(target_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        for line in lines:
            line = line.strip()
            if not line or any(line.lower().startswith(k) for k in ["date", "timestamp", "quarter", "observation_date"]):
                continue

            parts = [p.strip() for p in line.split(",")]
            if len(parts) >= 2:
                parsed_date = parse_quarter_label(parts[0])
                if not parsed_date:
                    continue
                try:
                    val = float(parts[1])
                    rev = parts[2] if len(parts) > 2 and parts[2] in ["ORIGINAL", "REVISED"] else "ORIGINAL"
                    observations.append({
                        "observation_date": parsed_date,
                        "value": val,
                        "unit": "tonnes",
                        "frequency": "quarterly",
                        "revision_status": rev,
                    })
                except ValueError:
                    continue

        return observations

    def parse_seed_data(self) -> List[Dict[str, Any]]:
        """Parses default seed dataset. Useful for isolated component tests."""
        raw_obs = self.parse_seed_csv(self.seed_path)
        return self.validate_and_apply_revisions(raw_obs)

    def extract_from_workbook(
        self, file_path: Union[str, Path], sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Extracts recycling observations from Excel workbooks or CSV sources.
        Propagates exceptions cleanly instead of suppressing them.
        """
        path = Path(file_path)
        observations: List[Dict[str, Any]] = []

        if not path.exists():
            raise FileNotFoundError(f"Source file not found: {path}")

        if path.suffix.lower() in [".csv", ".txt"]:
            return self.parse_seed_csv(path)

        if path.suffix.lower() in [".xlsx", ".xlsm"]:
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            sheet_candidates = ["Gold Balance", "Supply", "Supply_and_Demand"]
            sheet = next((name for name in sheet_candidates if name in wb.sheetnames), None)
            if sheet is None:
                raise ValueError(f"Target sheet not found. Available: {wb.sheetnames}")
            ws = wb[sheet]

            header_row = None
            quarter_columns = {}
            for row_idx in range(1, min(ws.max_row, 10) + 1):
                found = {}
                for col_idx in range(1, ws.max_column + 1):
                    parsed = parse_quarter_label(ws.cell(row_idx, col_idx).value)
                    if parsed:
                        found[col_idx] = parsed
                if found:
                    header_row, quarter_columns = row_idx, found
                    break
            if header_row is None:
                raise ValueError("Could not detect quarterly header row in workbook")

            matches = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                label = " ".join(
                    str(ws.cell(row_idx, col_idx).value or "")
                    for col_idx in (1, 2)
                ).strip().lower()
                if "recycled gold" in label or "recycling" in label:
                    matches.append(row_idx)
            if len(matches) > 1:
                raise ValueError("Ambiguous matching: Multiple recycled-gold rows detected")
            if len(matches) == 0:
                raise ValueError("Could not find recycled gold row in workbook")

            observations = []
            for col_idx, obs_date in quarter_columns.items():
                raw_value = ws.cell(matches[0], col_idx).value
                if raw_value is None or str(raw_value).strip() == "":
                    continue
                try:
                    value = float(raw_value)
                except (TypeError, ValueError) as exc:
                    raise ValueError(f"Non-numeric recycling value for {obs_date}: {raw_value}") from exc
                observations.append({
                    "observation_date": obs_date,
                    "value": value,
                    "unit": "tonnes",
                    "frequency": "quarterly",
                    "revision_status": "ORIGINAL",
                })
            if observations:
                return self.validate_and_apply_revisions(observations)

            # Deterministic fallback layout extraction
            import pandas as pd
            frame = pd.read_excel(path, sheet_name="Gold Balance", header=None)
            header_idx = next(
                (idx for idx, row in frame.iterrows()
                 if any(parse_quarter_label(value) for value in row.tolist())),
                None,
            )
            if header_idx is None:
                raise ValueError("Could not detect quarterly header row in Gold Balance")
            labels = frame.iloc[:, :2].fillna("").astype(str).agg(" ".join, axis=1).str.lower()
            matches = labels[labels.str.contains("recycled gold|recycling", regex=True)].index.tolist()
            if len(matches) != 1:
                raise ValueError("Ambiguous matching: expected exactly one recycled-gold row")
            row_idx = matches[0]
            fallback_observations = []
            for col_idx, value in enumerate(frame.iloc[header_idx].tolist()):
                obs_date = parse_quarter_label(value)
                if not obs_date:
                    continue
                raw_value = frame.iloc[row_idx, col_idx]
                if pd.isna(raw_value):
                    continue
                try:
                    numeric_value = float(raw_value)
                except (TypeError, ValueError) as exc:
                    raise ValueError(f"Non-numeric recycling value for {obs_date}: {raw_value}") from exc
                fallback_observations.append({
                    "observation_date": obs_date,
                    "value": numeric_value,
                    "unit": "tonnes",
                    "frequency": "quarterly",
                    "revision_status": "ORIGINAL",
                })
            return self.validate_and_apply_revisions(fallback_observations)

        # Generic pandas fallback path
        import pandas as pd
        df = pd.read_excel(path, sheet_name=sheet_name or 0)
        
        for idx, row in df.iterrows():
            row_vals = [v for v in row.values if pd.notna(v)]
            if len(row_vals) < 2:
                continue

            parsed_date = parse_quarter_label(str(row_vals[0]))
            if not parsed_date:
                parsed_date = parse_quarter_label(str(row_vals[1]))
                if not parsed_date:
                    continue
                value_idx = 2 if len(row_vals) > 2 else 1
            else:
                value_idx = 1

            val = float(row_vals[value_idx])
            observations.append({
                "observation_date": parsed_date,
                "value": val,
                "unit": "tonnes",
                "frequency": "quarterly",
                "revision_status": "ORIGINAL",
            })

        return self.validate_and_apply_revisions(observations)

    def run(
        self,
        publication_date: Optional[str] = None,
        source_file: Optional[Union[str, Path]] = None,
        is_live_source: bool = False,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Executes processing and writes schema-compliant output JSON."""
        retrieved_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        if not publication_date:
            publication_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

        target_file = Path(source_file) if source_file else self.seed_path
        using_seed_fallback = (target_file == self.seed_path) and not is_live_source

        raw_observations = self.extract_from_workbook(target_file)
        data = self.validate_and_apply_revisions(raw_observations, self.output_path)
        
        if using_seed_fallback:
            availability_status = "STALE"
        else:
            availability_status = "AVAILABLE" if data else "STALE"

        cached_file_hash = compute_sha256(target_file)

        payload = {
            "variable_id": self.variable_id,
            "availability_status": availability_status,
            "ingestion_metadata": {
                "publication_date": publication_date,
                "retrieved_at": retrieved_at,
                "source_file": str(target_file),
                "cached_file_hash": cached_file_hash,
            },
            "observations": data,
            "data": data,
        }

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)

        return payload


if __name__ == "__main__":
    collector = GoldRecyclingCollector()
    collector.run()