"""
parse_bar_and_coin.py — L0-005 parser
Extracts bar-and-coin demand data from WGC Gold Demand Trends quarterly workbook.

All paths are resolved relative to this script's directory.

Usage (from any working directory):
    python docs/phase2-ingestion/L0/005/data/parse_bar_and_coin.py \\
        --workbook gold-demand-trends/GDT_Tables_Q2\\'26_EN.xlsx \\
        --config config.yaml \\
        --publication-date 2026-08-07 \\
        --download-date 2026-08-18

Output:
    processed/L0_005_observations.csv   (relative to script directory)
    archive/ingest.log                  (relative to script directory)
"""

import argparse
import csv
import hashlib
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import yaml

PARSER_VERSION = "1.0.0"
VARIABLE_ID = "L0-005"
UNDERLYING_PROVIDERS = (
    "Metals Focus; Refinitiv GFMS; ICE Benchmark Administration; World Gold Council"
)

ANNUAL_PATTERN = re.compile(r"^\d{4}$")
QUARTERLY_PATTERN = re.compile(r"^Q([1-4])'(\d{2})$")

# All paths resolved relative to this script's directory
SCRIPT_DIR = Path(__file__).parent


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def setup_logging(log_path: Path) -> logging.Logger:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("L0-005")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s UTC | %(levelname)s | %(message)s")
    fmt.converter = lambda *_: datetime.now(timezone.utc).timetuple()
    if not logger.handlers:
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setFormatter(fmt)
        ch = logging.StreamHandler(sys.stdout)
        ch.setFormatter(fmt)
        logger.addHandler(fh)
        logger.addHandler(ch)
    return logger


# ---------------------------------------------------------------------------
# SHA-256
# ---------------------------------------------------------------------------

def sha256_of_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def load_workbook(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def detect_period_columns(ws, header_row: int) -> dict:
    """Returns {period_label: col_index} detected dynamically from header row."""
    cols = {}
    headers = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    for idx, val in enumerate(headers):
        if val is None:
            continue
        label = str(int(val)) if isinstance(val, (int, float)) and float(val) == int(val) else str(val).strip()
        if ANNUAL_PATTERN.match(label) or QUARTERLY_PATTERN.match(label):
            cols[label] = idx
    return cols


def get_row_values(ws, row_num: int) -> list:
    rows = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    if not rows:
        raise ValueError(f"Row {row_num} not found in sheet '{ws.title}'")
    return list(rows[0])


# ---------------------------------------------------------------------------
# Period parsing
# ---------------------------------------------------------------------------

def parse_period(label: str):
    """Returns (year: int, quarter: int|None, period_type: str)."""
    if ANNUAL_PATTERN.match(label):
        return int(label), None, "annual"
    m = QUARTERLY_PATTERN.match(label)
    if m:
        return 2000 + int(m.group(2)), int(m.group(1)), "quarterly"
    raise ValueError(f"Unrecognised period label: '{label}'")


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def extract_gold_balance(wb: openpyxl.Workbook, cfg: dict, _unused: dict) -> dict:
    sheet_cfg = cfg["extraction"]["gold_balance_sheet"]
    sheet_name = sheet_cfg["sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    cols = detect_period_columns(ws, sheet_cfg["header_row"])
    t = sheet_cfg["targets"]

    rows = {
        "total_bar_and_coin_tonnes":    get_row_values(ws, t["total_bar_and_coin"]["row"]),
        "bar_demand_tonnes":            get_row_values(ws, t["bars"]["row"]),
        "official_coin_demand_tonnes":  get_row_values(ws, t["official_coins"]["row"]),
        "medals_imitation_coin_tonnes": get_row_values(ws, t["medals_imitation_coins"]["row"]),
    }

    result = {}
    for label, col_idx in cols.items():
        _, _, ptype = parse_period(label)
        rec = {"total_bar_and_coin_tonnes": rows["total_bar_and_coin_tonnes"][col_idx]}
        if ptype == "annual":
            rec["bar_demand_tonnes"]            = rows["bar_demand_tonnes"][col_idx]
            rec["official_coin_demand_tonnes"]  = rows["official_coin_demand_tonnes"][col_idx]
            rec["medals_imitation_coin_tonnes"] = rows["medals_imitation_coin_tonnes"][col_idx]
        else:
            rec["bar_demand_tonnes"]            = None
            rec["official_coin_demand_tonnes"]  = None
            rec["medals_imitation_coin_tonnes"] = None
        result[label] = rec
    return result


def extract_bar_and_coin_sheet(wb: openpyxl.Workbook, cfg: dict) -> dict:
    sheet_cfg = cfg["extraction"]["bar_and_coin_sheet"]
    sheet_name = sheet_cfg["sheet_name"]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    cols = detect_period_columns(ws, sheet_cfg["header_row"])
    t = sheet_cfg["targets"]

    row_named = get_row_values(ws, t["named_country_total"]["row"])
    row_other = get_row_values(ws, t["other_and_stock_change"]["row"])
    row_world = get_row_values(ws, t["world_total"]["row"])

    return {
        label: {
            "named_country_total_tonnes":            row_named[col_idx],
            "other_and_stock_change_tonnes":         row_other[col_idx],
            "world_total_bar_and_coin_sheet_tonnes": row_world[col_idx],
        }
        for label, col_idx in cols.items()
    }


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_record(rec: dict, cfg: dict, logger: logging.Logger) -> tuple:
    vcfg = cfg["validation"]
    notes = []
    status = "PASS"
    total = rec.get("total_bar_and_coin_tonnes")
    ptype = rec["observation_period_type"]

    if total is None or total < vcfg["global_total_min"]:
        return "FAIL", f"total_bar_and_coin_tonnes={total} is None or negative"

    rng = vcfg["annual_plausible_range"] if ptype == "annual" else vcfg["quarterly_plausible_range"]
    if not (rng["min"] <= total <= rng["max"]):
        notes.append(f"total={total:.5f} outside plausible {ptype} range [{rng['min']}, {rng['max']}]t")
        status = "FLAG"

    if ptype == "annual":
        bar    = rec.get("bar_demand_tonnes")
        coins  = rec.get("official_coin_demand_tonnes")
        medals = rec.get("medals_imitation_coin_tonnes")
        if all(v is not None for v in [bar, coins, medals]):
            comp_sum = bar + coins + medals
            diff_pct = abs(comp_sum - total) / total * 100 if total else 0
            if diff_pct > vcfg["subcomponent_fail_threshold_pct"]:
                return "FAIL", (
                    f"Sub-component sum {comp_sum:.5f} differs from total {total:.5f} "
                    f"by {diff_pct:.2f}% (FAIL threshold {vcfg['subcomponent_fail_threshold_pct']}%)"
                )
            if diff_pct > vcfg["subcomponent_tolerance_pct"]:
                notes.append(f"Sub-component sum {comp_sum:.5f} differs from total by {diff_pct:.2f}%")
                status = "FLAG"

    world = rec.get("world_total_bar_and_coin_sheet_tonnes")
    if world is not None and total:
        diff_pct = abs(world - total) / total * 100
        if diff_pct > vcfg["sheet_reconciliation_tolerance_pct"]:
            notes.append(
                f"Bar and Coin sheet world_total={world:.5f} differs from "
                f"Gold Balance total={total:.5f} by {diff_pct:.4f}%"
            )
            status = "FLAG"

    label = rec["observation_period"]
    if not (ANNUAL_PATTERN.match(label) or QUARTERLY_PATTERN.match(label)):
        return "FAIL", f"Period label '{label}' does not match expected format"

    return status, ("; ".join(notes) if notes else None)


# ---------------------------------------------------------------------------
# Revision detection
# ---------------------------------------------------------------------------

def load_existing_records(processed_path: Path) -> dict:
    existing = {}
    if not processed_path.exists():
        return existing
    with open(processed_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            period = row.get("observation_period")
            total  = row.get("total_bar_and_coin_tonnes")
            if period and total:
                try:
                    existing[period] = float(total)
                except ValueError:
                    pass
    return existing


# ---------------------------------------------------------------------------
# Record assembly
# ---------------------------------------------------------------------------

def assemble_records(gb_data, bc_data, workbook_name, sha256,
                     publication_date, download_date, ingested_at,
                     existing, cfg, logger) -> list:
    records = []
    all_periods = sorted(
        set(gb_data.keys()) | set(bc_data.keys()),
        key=lambda p: (parse_period(p)[0], parse_period(p)[1] or 0),
    )
    for label in all_periods:
        if label not in gb_data:
            logger.warning(f"Period {label} in Bar and Coin sheet but not Gold Balance — skipping")
            continue

        year, quarter, ptype = parse_period(label)
        gb = gb_data[label]
        bc = bc_data.get(label, {})

        total = gb["total_bar_and_coin_tonnes"]
        prior_total = existing.get(label)
        is_revised = (
            prior_total is not None and
            total is not None and
            abs(float(total) - float(prior_total)) > 1e-6
        )

        rec = {
            "variable_id":                          VARIABLE_ID,
            "observation_period":                   label,
            "observation_period_type":              ptype,
            "observation_year":                     year,
            "observation_quarter":                  quarter,
            "bar_demand_tonnes":                    gb.get("bar_demand_tonnes"),
            "official_coin_demand_tonnes":          gb.get("official_coin_demand_tonnes"),
            "medals_imitation_coin_tonnes":         gb.get("medals_imitation_coin_tonnes"),
            "total_bar_and_coin_tonnes":            total,
            "named_country_total_tonnes":           bc.get("named_country_total_tonnes"),
            "other_and_stock_change_tonnes":        bc.get("other_and_stock_change_tonnes"),
            "world_total_bar_and_coin_sheet_tonnes": bc.get("world_total_bar_and_coin_sheet_tonnes"),
            "unit":                                 "metric_tonnes",
            "source_name":                          "WGC_GDT",
            "source_workbook":                      workbook_name,
            "source_publication_date":              publication_date,
            "download_date":                        download_date,
            "workbook_sha256":                      sha256,
            "ingested_at":                          ingested_at,
            "parser_version":                       PARSER_VERSION,
            "underlying_providers":                 UNDERLYING_PROVIDERS,
            "is_revised":                           is_revised,
            "prior_workbook_sha256":                "see_prior_record" if is_revised else None,
            "prior_total_bar_and_coin_tonnes":      prior_total if is_revised else None,
            "revision_reason": (
                f"Value changed from {prior_total:.5f}t to {total:.5f}t in {workbook_name}"
                if is_revised else None
            ),
            "validation_status":  None,
            "availability_status": None,
            "anomaly_notes":       None,
        }

        status, notes = validate_record(rec, cfg, logger)
        rec["validation_status"]  = status
        rec["availability_status"] = "AVAILABLE" if status in ("PASS", "FLAG") else "INCOMPLETE"
        rec["anomaly_notes"]       = notes
        records.append(rec)
    return records


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

FIELDNAMES = [
    "variable_id", "observation_period", "observation_period_type",
    "observation_year", "observation_quarter",
    "bar_demand_tonnes", "official_coin_demand_tonnes", "medals_imitation_coin_tonnes",
    "total_bar_and_coin_tonnes",
    "named_country_total_tonnes", "other_and_stock_change_tonnes",
    "world_total_bar_and_coin_sheet_tonnes",
    "unit", "source_name", "source_workbook",
    "source_publication_date", "download_date", "workbook_sha256",
    "ingested_at", "parser_version", "underlying_providers",
    "is_revised", "prior_workbook_sha256", "prior_total_bar_and_coin_tonnes",
    "revision_reason", "validation_status", "availability_status", "anomaly_notes",
]


def append_to_csv(records: list, processed_path: Path):
    processed_path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not processed_path.exists()
    with open(processed_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, extrasaction="ignore")
        if write_header:
            writer.writeheader()
        writer.writerows(records)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="L0-005 Bar-and-Coin parser")
    ap.add_argument("--workbook",         required=True)
    ap.add_argument("--config",           required=True)
    ap.add_argument("--publication-date", required=True)
    ap.add_argument("--download-date",    required=True)
    ap.add_argument("--dry-run",          action="store_true")
    args = ap.parse_args()

    config_path = SCRIPT_DIR / args.config
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    log_path       = SCRIPT_DIR / "archive" / "ingest.log"
    processed_path = SCRIPT_DIR / "processed" / "L0_005_observations.csv"

    logger = setup_logging(log_path)

    workbook_path = SCRIPT_DIR / args.workbook
    if not workbook_path.exists():
        logger.error(f"Workbook not found: {workbook_path}")
        sys.exit(1)

    logger.info(f"Workbook: {workbook_path.name}")
    sha256 = sha256_of_file(workbook_path)
    logger.info(f"SHA-256: {sha256}")

    ingested_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    existing    = load_existing_records(processed_path)
    logger.info(f"Existing periods in processed store: {len(existing)}")

    wb      = load_workbook(workbook_path)
    logger.info("Extracting Gold Balance sheet...")
    gb_data = extract_gold_balance(wb, cfg, {})
    n_ann = sum(1 for l in gb_data if ANNUAL_PATTERN.match(l))
    n_qtr = sum(1 for l in gb_data if QUARTERLY_PATTERN.match(l))
    logger.info(f"Gold Balance: {n_ann} annual, {n_qtr} quarterly periods")

    logger.info("Extracting Bar and Coin sheet...")
    bc_data = extract_bar_and_coin_sheet(wb, cfg)

    records = assemble_records(
        gb_data, bc_data, workbook_path.name, sha256,
        args.publication_date, args.download_date, ingested_at,
        existing, cfg, logger,
    )

    n_pass = sum(1 for r in records if r["validation_status"] == "PASS")
    n_flag = sum(1 for r in records if r["validation_status"] == "FLAG")
    n_fail = sum(1 for r in records if r["validation_status"] == "FAIL")
    n_rev  = sum(1 for r in records if r["is_revised"])
    n_ann2 = sum(1 for r in records if r["observation_period_type"] == "annual")
    n_qtr2 = sum(1 for r in records if r["observation_period_type"] == "quarterly")

    logger.info(f"Records assembled: {len(records)} total ({n_ann2} annual, {n_qtr2} quarterly)")
    logger.info(f"Validation: {n_pass} PASS, {n_flag} FLAG, {n_fail} FAIL")
    logger.info(f"Revisions detected: {n_rev}")

    if n_fail > 0:
        for r in records:
            if r["validation_status"] == "FAIL":
                logger.error(f"FAIL — {r['observation_period']}: {r['anomaly_notes']}")
        logger.error("FAIL records present. Output not written.")
        sys.exit(2)

    if n_flag > 0:
        for r in records:
            if r["validation_status"] == "FLAG":
                logger.warning(f"FLAG — {r['observation_period']}: {r['anomaly_notes']}")
        logger.warning("FLAG records require operator review before scoring.")

    if args.dry_run:
        logger.info("Dry run — no output written.")
        return

    append_to_csv(records, processed_path)
    logger.info(f"Appended {len(records)} records to {processed_path}")


if __name__ == "__main__":
    main()
