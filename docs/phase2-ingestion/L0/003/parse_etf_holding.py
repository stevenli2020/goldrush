import csv
import json
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

class GoldETFHoldingsParser:

    def __init__(self, parser_version='1.0.0', log_path='archive/ingest.log'):
        self.parser_version = parser_version
        self.log_path = Path(log_path)
        self.validation_errors = []

    def calculate_metadata(self, file_path: Path) -> str:
        return None

    def _log_error(self, row_idx, raw_data, reason, action='REJECTED'):
        entry = {'timestamp': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'), 'row_index': row_idx, 'raw_data': str(raw_data), 'reason': reason, 'action': action}
        self.validation_errors.append(entry)
        self.log_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.log_path, 'a', encoding='utf-8') as f:
            f.write(json.dumps(entry) + '\n')

    def parse_file(self, file_path, publication_date, download_date, prior_records=None):
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f'Source file not found: {path}')
        if path.suffix.lower() in ['.xlsx', '.xls']:
            return self.parse_workbook(path, publication_date, download_date, prior_records)
        else:
            return self.parse_csv(path, publication_date, download_date, prior_records)

    def parse_workbook(self, file_path: Path, publication_date: str, download_date: str, prior_records=None):
        file_metadata = None
        ingestion_ts = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
        prior_records = prior_records or {}
        wb = openpyxl.load_workbook(file_path, data_only=True)
        target_sheet = None
        for sheet in ['Holdings by month', 'Holdings', 'Historical Data', wb.sheetnames[0]]:
            if sheet in wb.sheetnames:
                target_sheet = wb[sheet]
                break
        rows = list(target_sheet.iter_rows(values_only=True))
        header_idx = None
        for idx, row in enumerate(rows):
            row_str = [str(c).strip().lower() for c in row if c is not None]
            if any(('date' in c for c in row_str)) and any(('holding' in c or 'tonnes' in c for c in row_str)):
                header_idx = idx
                break
        if header_idx is None:
            raise ValueError(f'Could not locate valid header row in workbook {file_path}')
        headers = [str(c).strip() if c is not None else '' for c in rows[header_idx]]
        date_col = next((i for i, h in enumerate(headers) if 'date' in h.lower()), 0)
        region_col = next((i for i, h in enumerate(headers) if 'region' in h.lower()), None)
        tonnes_col = next((i for i, h in enumerate(headers) if 'holding' in h.lower() or 'tonnes' in h.lower()), 1)
        aum_col = next((i for i, h in enumerate(headers) if 'aum' in h.lower()), None)
        records = []
        seen_keys = set()
        prev_value = None
        for row_num, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not any(row):
                continue
            raw_date = row[date_col]
            if raw_date is None:
                continue
            if isinstance(raw_date, datetime):
                obs_date = raw_date.strftime('%Y-%m-%d')
            else:
                try:
                    obs_date = datetime.strptime(str(raw_date).strip(), '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    self._log_error(row_num, row, f'Invalid observation date format: {raw_date}')
                    continue
            region = str(row[region_col]).strip() if region_col is not None and row[region_col] else 'GLOBAL'
            key = (obs_date, region)
            if key in seen_keys:
                self._log_error(row_num, row, f'Duplicate observation key: {key}', action='SKIPPED')
                continue
            seen_keys.add(key)
            try:
                tonnes_val = float(row[tonnes_col])
            except (ValueError, TypeError):
                self._log_error(row_num, row, f'Non-numeric holdings tonnes value: {row[tonnes_col]}')
                continue
            if tonnes_val < 0:
                self._log_error(row_num, row, f'Negative holdings tonnes value: {tonnes_val}')
                continue
            aum_val = None
            if aum_col is not None and row[aum_col] is not None:
                try:
                    aum_val = float(row[aum_col])
                except (ValueError, TypeError):
                    aum_val = None
            val_status = 'PASS'
            if prev_value is not None and prev_value > 0:
                pct_change = abs(tonnes_val - prev_value) / prev_value
                if pct_change > 0.05:
                    val_status = 'FLAG'
            prev_value = tonnes_val
            prior_val = prior_records.get(key)
            is_revision = prior_val is not None and prior_val != tonnes_val
            record = {'observation_date': obs_date, 'region': region, 'holdings_tonnes': round(tonnes_val, 2), 'aum_usd_bn': round(aum_val, 2) if aum_val is not None else None, 'unit': 'metric_tonnes', 'source_citation': 'World Gold Council', 'source_file': file_path.name, 'publication_date': publication_date, 'download_date': download_date, 'ingestion_timestamp': ingestion_ts, 'parser_version': self.parser_version, 'revision_metadata': {'is_revision': is_revision, 'prior_value': prior_val, 'revision_reason': 'Value updated from baseline' if is_revision else None}, 'validation_status': val_status, 'availability_status': 'AVAILABLE'}
            records.append(record)
        return records

    def parse_csv(self, file_path: Path, publication_date: str, download_date: str, prior_records=None):
        file_metadata = None
        ingestion_ts = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
        prior_records = prior_records or {}
        records = []
        seen_keys = set()
        prev_value = None
        with open(file_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):
                obs_date = row.get('observation_date', '').strip()
                region = row.get('region', 'GLOBAL').strip()
                val_str = row.get('holdings_tonnes', '').strip()
                aum_str = row.get('aum_usd_bn', '').strip()
                try:
                    datetime.strptime(obs_date, '%Y-%m-%d')
                except ValueError:
                    self._log_error(row_num, row, f'Invalid date: {obs_date}')
                    continue
                key = (obs_date, region)
                if key in seen_keys:
                    self._log_error(row_num, row, f'Duplicate date: {obs_date}', action='SKIPPED')
                    continue
                seen_keys.add(key)
                try:
                    val = float(val_str)
                except ValueError:
                    self._log_error(row_num, row, f'Non-numeric value: {val_str}')
                    continue
                if val < 0:
                    self._log_error(row_num, row, f'Negative value: {val}')
                    continue
                aum_val = float(aum_str) if aum_str else None
                val_status = 'PASS'
                if prev_value is not None and prev_value > 0:
                    if abs(val - prev_value) / prev_value > 0.05:
                        val_status = 'FLAG'
                prev_value = val
                prior_val = prior_records.get(key)
                is_revision = prior_val is not None and prior_val != val
                records.append({'observation_date': obs_date, 'region': region, 'holdings_tonnes': round(val, 2), 'aum_usd_bn': round(aum_val, 2) if aum_val is not None else None, 'unit': 'metric_tonnes', 'source_citation': 'World Gold Council', 'source_file': file_path.name, 'publication_date': publication_date, 'download_date': download_date, 'ingestion_timestamp': ingestion_ts, 'parser_version': self.parser_version, 'revision_metadata': {'is_revision': is_revision, 'prior_value': prior_val, 'revision_reason': 'Value updated from baseline' if is_revision else None}, 'validation_status': val_status, 'availability_status': 'AVAILABLE'})
        return records

    def generate_stale_fallback(self, base_record: dict, target_date: str) -> dict:
        stale_record = base_record.copy()
        stale_record['observation_date'] = target_date
        stale_record['availability_status'] = 'STALE'
        return stale_record
